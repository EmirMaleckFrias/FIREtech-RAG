"""Parseo de documentos: el único camino de ingesta del sistema.

Contrato (SPEC.md, "Gestión de documentos"):
- PDF   → texto por página (pdfplumber), párrafos RECONSTRUIDOS a partir de
          las líneas físicas (ver `_abre_parrafo`), chunks de ~400 tokens con
          overlap del 15%; `page` = primera página del chunk, `source_pages` =
          todas, `chunk_type` = "text".
- DOCX  → párrafos agrupados por sección (el encabezado vigente) y tablas en
          bloques de filas que repiten la cabecera; Word no tiene páginas, así
          que se cita por sección (texto) o por número de tabla.
- XLSX/CSV → detección de fila de encabezado, un chunk por fila
          ("Campo: valor"); `chunk_type` = "table", `page` = número de fila.
- TXT/MD → chunks por párrafos; `page` = índice de chunk (1-based).

Los chunks de PDF y DOCX llevan ANTEPUESTA la sección vigente, y los de PDF
además el título de la obra (ver `_con_contexto`): el vector se calcula sobre
`text` y nada más, así que el contexto tiene que viajar dentro del texto
embebido y no solo en el payload, que nadie ve hasta después de recuperar.

Cada chunk producido es un dict con TODAS las claves de
`app.services.qdrant._PAYLOAD_KEYS` + `id` (uuid4), listo para
embed_texts → upsert_chunks. `project_id` y `document_id` los rellena quien
llama, que es el único que sabe a qué proyecto pertenece el archivo.

Saneo: chunks con texto vacío se omiten; documentos que generan más de
MAX_CHUNKS chunks (o cero) levantan ValueError con mensaje claro.
"""
from __future__ import annotations

import csv
import io
import logging
import re
import uuid
from pathlib import Path
from typing import TypeVar

from app.ingest.idioma import detectar_idioma

logger = logging.getLogger(__name__)

MAX_CHUNKS = 4000          # tope duro por documento (error claro si se excede)
_TARGET_TOKENS = 400       # tamaño objetivo de chunk (aprox.)
_OVERLAP_TOKENS = 60       # 15% de 400
_MAX_PARA_TOKENS = 500     # párrafos más largos se subdividen por oraciones
_MAX_CHUNK_CHARS = 8000    # límite duro de texto por chunk (= truncado embeddings)

# Localizador que acompaña a cada párrafo al empaquetar: la página (PDF), las
# páginas que cruza, o nada (DOCX, TXT). El empaquetador no lo mira.
_L = TypeVar("_L")

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _est_tokens(text: str) -> int:
    """Estimación barata de tokens (~4 chars/token para es/en)."""
    return max(1, len(text) // 4)


def _base_chunk(
    file_name: str,
    text: str,
    page: int,
    source_pages: list[int],
    chunk_type: str,
    section: str = "",
    source_row: int | None = None,
    meta: object | None = None,
) -> dict:
    """Dict de chunk con TODAS las claves de qdrant._PAYLOAD_KEYS + id.

    `project_id` y `document_id` los rellena quien llama (la ruta de subida o
    el CLI de ingesta), porque el parser no sabe a qué proyecto pertenece el
    archivo. `language` queda vacío hasta que se detecte.

    `meta` es un `paper.PaperMeta` cuando el documento es un artículo: lo que
    permite citar "Allegri et al., 2023" en vez del nombre del archivo. Se
    repite en cada fragmento a propósito, para que una cita no necesite ir a
    buscar nada más.
    """
    text = text[:_MAX_CHUNK_CHARS]
    return {
        "id": str(uuid.uuid4()),
        "text": text,
        "source_file": file_name,
        "page": page,
        "project_id": None,
        "document_id": None,
        "document_version": None,
        "section": section,
        "language": "",
        "document_type": Path(file_name).suffix.lower().lstrip(".") or "unknown",
        "source_pages": source_pages,
        "metadata": {"source_row": source_row} if source_row is not None else {},
        "chunk_type": chunk_type,
        "title": getattr(meta, "titulo", "") or "",
        "citation": getattr(meta, "referencia", "") or "",
        "doi": getattr(meta, "doi", "") or "",
    }


def _split_long_paragraph(para: str) -> list[str]:
    """Subdivide párrafos que exceden _MAX_PARA_TOKENS (por oraciones;
    si una 'oración' sigue siendo enorme (texto sin puntuación), por palabras)."""
    if _est_tokens(para) <= _MAX_PARA_TOKENS:
        return [para]

    sentences: list[str] = []
    for sent in re.split(r"(?<=[.!?;:])\s+", para):
        sent = sent.strip()
        if not sent:
            continue
        if _est_tokens(sent) <= _MAX_PARA_TOKENS:
            sentences.append(sent)
        else:  # sin puntuación: corte duro por palabras
            words = sent.split()
            cur: list[str] = []
            cur_len = 0
            for word in words:
                cur.append(word)
                cur_len += len(word) + 1
                if cur_len // 4 >= _TARGET_TOKENS:
                    sentences.append(" ".join(cur))
                    cur, cur_len = [], 0
            if cur:
                sentences.append(" ".join(cur))

    # Empaqueta oraciones en piezas de ~_TARGET_TOKENS.
    pieces: list[str] = []
    cur_sents: list[str] = []
    cur_tok = 0
    for sent in sentences:
        stok = _est_tokens(sent)
        if cur_sents and cur_tok + stok > _TARGET_TOKENS:
            pieces.append(" ".join(cur_sents))
            cur_sents, cur_tok = [], 0
        cur_sents.append(sent)
        cur_tok += stok
    if cur_sents:
        pieces.append(" ".join(cur_sents))
    return pieces


def _split_paragraphs(text: str) -> list[str]:
    """Párrafos por líneas en blanco; los muy largos se subdividen."""
    out: list[str] = []
    for para in re.split(r"\n\s*\n", text):
        para = para.strip()
        if para:
            out.extend(_split_long_paragraph(para))
    return out


def _pack_paragraphs(
    paras: list[tuple[str, _L]]
) -> list[list[tuple[str, _L]]]:
    """Agrupa (párrafo, localizador) en chunks de ~_TARGET_TOKENS con overlap
    de ~_OVERLAP_TOKENS tomado de los párrafos finales del chunk anterior.

    Es el ÚNICO empaquetador del módulo, y PDF, DOCX y TXT pasan por él. Hasta
    sep 2026 solo lo usaba TXT: PDF y DOCX tenían bucles propios sin solape,
    contradiciendo el contrato del módulo, y una oración que caía en la
    frontera de dos chunks no estaba entera en ninguno. El localizador viaja
    opaco: solo el texto cuenta tokens.
    """
    chunks: list[list[tuple[str, _L]]] = []
    cur: list[tuple[str, _L]] = []
    cur_tok = 0
    for para, page in paras:
        ptok = _est_tokens(para)
        if cur and cur_tok + ptok > _TARGET_TOKENS:
            chunks.append(cur)
            # Overlap: párrafos finales hasta cubrir ~_OVERLAP_TOKENS.
            tail: list[tuple[str, _L]] = []
            ttok = 0
            for prev in reversed(cur):
                t = _est_tokens(prev[0])
                if tail and ttok + t > _OVERLAP_TOKENS:
                    break
                tail.insert(0, prev)
                ttok += t
                if ttok >= _OVERLAP_TOKENS:
                    break
            # Si el overlap fuese el chunk entero no aporta nada (y duplicaría
            # todo el texto): se descarta.
            if len(tail) == len(cur):
                tail = []
            cur = list(tail)
            cur_tok = sum(_est_tokens(p) for p, _ in cur)
        cur.append((para, page))
        cur_tok += ptok
    if cur:
        chunks.append(cur)
    return chunks


def _agrupar_por_seccion(
    paras: list[tuple[str, _L, str]]
) -> list[tuple[str, list[tuple[str, _L]]]]:
    """Tramos consecutivos de una misma sección: [(sección, [(texto, loc)])].

    Se empaqueta tramo a tramo, sin mezclar secciones, para que la cita de un
    fragmento apunte a una sección de verdad y no a la frontera entre dos. El
    solape tampoco cruza secciones a propósito: un dato de Resultados no debe
    aparecer "de cola" en el primer chunk de Discusión, donde se leería como
    interpretación del autor.
    """
    grupos: list[tuple[str, list[tuple[str, _L]]]] = []
    for texto, loc, sec in paras:
        if not grupos or grupos[-1][0] != sec:
            grupos.append((sec, []))
        grupos[-1][1].append((texto, loc))
    return grupos


def _con_contexto(cuerpo: str, *contexto: str) -> str:
    """Antepone al texto de un chunk sus líneas de contexto (título, sección).

    Por qué: el embedding se calcula SOLO sobre `text`. Un fragmento de
    Resultados que dice "the mean was 542 pg/mL in the impaired group" no
    contiene ni la palabra Results ni de qué estudio sale; frente a la consulta
    "amyloid levels in early Alzheimer disease" puntúa peor que el mismo texto
    con "Cerebrospinal fluid biomarkers in early Alzheimer disease / Results"
    delante. Y cuando el LLM lee el fragmento sabe si está ante evidencia
    (Resultados) o interpretación (Discusión) sin mirar el payload.

    No se repite lo que ya está: si dos líneas de contexto coinciden (la
    portada de un PDF, donde el título se detecta como encabezado por formato y
    pasa a ser la sección) o si el cuerpo ya arranca con esa línea (Word indexa
    el encabezado como bloque propio del primer chunk), se deja una sola vez.
    """
    primera = re.sub(r"\s+", " ", cuerpo.split("\n", 1)[0]).strip().lower()
    lineas: list[str] = []
    for linea in contexto:
        linea = re.sub(r"\s+", " ", linea or "").strip()
        if not linea or linea.lower() == primera:
            continue
        if any(linea.lower() == previa.lower() for previa in lineas):
            continue
        lineas.append(linea)
    if not lineas:
        return cuerpo
    return "\n".join(lineas) + "\n\n" + cuerpo


# ---------------------------------------------------------------------------
# Reconstrucción de párrafos a partir de líneas físicas (PDF)
# ---------------------------------------------------------------------------
# Marcadores que abren párrafo por sí solos, aunque la línea anterior no haya
# cerrado la oración: viñetas y rótulos de tabla o figura. Un rótulo ("Table 1.
# Baseline characteristics") suele venir tras una fila de tabla que no acaba en
# punto, y pegarlo al párrafo anterior le quitaría su condición de rótulo. Las
# enumeraciones "(1)", "a)" NO cuentan: en un artículo van dentro de la oración
# ("two criteria: (1) age... and (2) ...") y una línea que empieza por "(2) and"
# es continuación. Los símbolos van como escape: el guion largo no debe
# aparecer literalmente en ningún archivo del proyecto (ver paper.normalizar).
_VINETA = re.compile(r"^[-*+\u2022\u00b7\u25e6\u25aa\u2023\u2013\u2014]\s+")
# El rótulo exige puntuación de rótulo tras el número ("Table 1." / "Figure 3:"
# / "TABLE 1 |"): sin ella, "as summarized in" | "Table 1, the groups differed"
# (una referencia dentro de la oración que cae a inicio de línea, cosa que pasa
# en casi todo párrafo de Resultados) se partiría por la mitad. Se pierde el
# rótulo sin puntuación ("Table 1 Baseline characteristics", estilo Springer),
# que entonces se pega a la fila anterior: no se pierde nada, solo se junta.
#
# La puntuación NO basta por sí sola, y por eso el rótulo solo abre párrafo si
# la línea anterior no está a mitad de oración (ver `_abre_parrafo`): medido el
# 4 sep 2026, `_abre_parrafo("summarized in", "Table 1. The groups did not
# differ.")` devolvía True y partía la oración en dos, porque el punto que
# sigue al número era el de "Table 1." pero cerraba la frase de la referencia.
_ROTULO = re.compile(
    r"^(?:table|tabla|cuadro|figure|figura|fig\.|box)\s*S?\d+[a-z]?\s*[.:|]",
    re.IGNORECASE,
)
# En Word el rótulo es un párrafo propio y no hay oración que partir, así que
# basta con que empiece por "Tabla N"; un falso positivo ("Table 1 shows...")
# solo añade al chunk de la tabla la frase que la describe.
_ROTULO_TABLA = re.compile(r"^(?:table|tabla|cuadro)\s*S?\d", re.IGNORECASE)
# Fin de oración, admitiendo el cierre de comillas o paréntesis tras el punto.
_FIN_DE_ORACION = re.compile(r"[.?!:][\)\]\"'\u201d\u2019]*$")
# Fin de oración con la cita Vancouver pegada al punto ("cognitive decline.12,13",
# "the first one.14", "decline.12-14"), que es el estilo de la mayoría de
# revistas médicas: el superíndice se extrae como texto normal, así que sin esta
# variante NINGUNA oración citada cerraba y el párrafo crecía hasta el corte
# duro de _MAX_PARA_TOKENS, en mitad de una frase. Medido el 4 sep 2026 con un
# PDF de tres oraciones citadas: salían como UN párrafo con la variante
# desactivada y como tres con ella.
#
# Exige una LETRA justo antes del punto, y ahí está la defensa de los decimales
# y las versiones: "0.31" y "gpt-5.4" llevan un dígito delante del punto, así
# que no cierran oración por esta vía y siguen sin partir la línea.
_FIN_CON_CITA = re.compile(
    r"[^\W\d_][.?!]\d{1,3}(?:\s*[,;\u2010\u2013-]\s*\d{1,3})*$"
)
# Palabra cortada por el maquetador: una letra seguida de guion al final.
_CORTE_DE_PALABRA = re.compile(r"[^\W\d_][-\u2010\u00ad]$")
# Cualquier guion pegado a una palabra o cifra al final ("COVID-", "12-").
_GUION_FINAL = re.compile(r"\w[-\u2010\u00ad\u2013]$")
# Prefijos que forman término con guion: si la palabra cortada acaba en uno de
# ellos, el guion es del compuesto y no del maquetador (ver `_unir_lineas`).
_PREFIJOS_CON_GUION = frozenset(
    "anti pre post non self co re sub inter intra multi beta alpha tau p e4".split()
)
# Además del prefijo conocido, cualquier izquierda de hasta tantos caracteres se
# considera compuesto ("long-term", "well-being", "p-value"): una palabra
# partida por el maquetador raramente deja una primera mitad tan corta.
_MAX_IZQUIERDA_CON_GUION = 5
# Abreviaturas con punto que NO cierran la oración. Corta a propósito: "etc."
# o "no." sí pueden cerrarla, y una lista larga inventa más de lo que arregla.
_ABREVIATURAS = (
    "et al.", "e.g.", "i.e.", "vs.", "cf.", "fig.", "figs.", "eq.", "ref.",
    "approx.", "ca.", "p. ej.", "vol.", "pp.",
)

# Palabras de enlace (preposiciones, conjunciones, artículos, auxiliares) en
# es/en. Son la señal de que una línea sigue a mitad de oración: ni una fila de
# tabla ni un subtítulo acaban en "and" o "de", y una línea de prosa cortada por
# el maquetador lo hace constantemente ("were assessed by" | "Smith and...").
_ENLACES = frozenset(
    """
    a about above after against all also among an and any are as at be because
    been before being below both but by during each either for from further had
    has have how however if in into is it its less many more most much neither
    no not of on only or other over per same several since so some such than
    that the their then there therefore these this those though through thus to
    under until upon very was were what when where whether which while with
    within without
    al ademas ante antes aunque como con cada cual cuando de del desde donde
    durante el en entonces entre era eran es fue fueron ha han hacia hasta la
    las los luego mas mientras muy para pero por pues que quien salvo se segun
    si sin sobre solo su sus tambien tanto tras un una unos unas varios y
    """.split()
)

# Señal de rejilla: dos o más huecos de >=2 espacios. Es la que dan los PDF de
# verdad, donde el maquetador separa las celdas con espacio ancho.
_HUECO_DE_REJILLA = re.compile(r"\S {2,}(?=\S)")
# Celda numérica: cifra con o sin decimales, rango "12-15", porcentaje. El
# adorno (paréntesis de la SD, "<" del p-valor, ± del error) se recorta antes.
_CELDA_NUMERICA = re.compile(
    r"[-+\u2212]?\d+(?:[.,]\d+)?(?:[-\u2013/][-+\u2212]?\d+(?:[.,]\d+)?)?%?"
)
_ADORNO_DE_CELDA = "()[]{}<>\u00ab\u00bb\"'\u201c\u201d.,;:*\u2020\u2021\u00a7"
# Un párrafo de PDF que aún tiene <= tantas palabras es demasiado corto para ser
# prosa: si además no acaba en puntuación, es un subtítulo que no se detectó.
_MAX_PALABRAS_SUELTA = 6


def _es_celda_numerica(token: str) -> bool:
    """¿El token es una celda de datos numérica? "(6.1)", "<0.001", "45%", "12-15"."""
    limpio = token.strip(_ADORNO_DE_CELDA).lstrip("<>=~\u2264\u2265\u00b1")
    if not any(caracter.isdigit() for caracter in limpio):
        return False
    return bool(_CELDA_NUMERICA.fullmatch(limpio))


def _parece_fila_de_tabla(linea: str) -> bool:
    """¿La línea es una FILA DE TABLA y no prosa? Decide si NO se puede pegar.

    Regresión medida el 4 sep 2026 contra el parser de líneas: al reconstruir
    párrafos con la regla "una línea que no cierra oración continúa el párrafo",
    una tabla de basales dentro de un PDF perdía las filas. "Table 1. Baseline
    characteristics / Variable Control MCI AD p / Age, years 72.4 (6.1) ... /
    MMSE 28.1 (1.2) ..." salía como UNA sola línea con todos los números
    seguidos: las filas no acaban en punto, así que se pegaban unas a otras y
    ningún número quedaba junto a su cabecera.

    Dos señales, ninguna de las cuales necesita mirar la geometría del PDF:

    1. Rejilla de espacios: >=2 huecos de dos o más espacios (lo que da un PDF
       maquetado de verdad; en los PDF de test las celdas van con un espacio y
       esta señal no dispara, de ahí que haga falta la segunda).
    2. Densidad numérica: >=2 celdas numéricas y, DESDE la primera de ellas, ni
       una palabra de enlace. Es lo que separa "Age, years 72.4 (6.1) 74.0 (5.8)
       0.31" (todo cifras tras la etiqueta) de la prosa con datos, que siempre
       las lleva: "Mean amyloid beta 42 was 542" ("was"), "recruited 120
       participants ... between 2018 and" ("between", "and"), "the coefficient
       was -0.31 (p = 0.02) in the adjusted model" ("in", "the").

    En un RAG médico las tablas de basales y de resultados son justo donde están
    los datos, así que ante la duda la regla se equivoca hacia NO pegar: cortar
    de más deja una fila como párrafo propio (que el empaquetador vuelve a
    juntar en el mismo chunk), y cortar de menos deja el número sin cabecera.
    """
    linea = linea.strip()
    if len(_HUECO_DE_REJILLA.findall(linea)) >= 2:
        return True
    tokens = linea.split()
    if len(tokens) < 2:
        return False
    numericos = [i for i, token in enumerate(tokens) if _es_celda_numerica(token)]
    if len(numericos) < 2:
        return False
    return not any(
        token.strip(_ADORNO_DE_CELDA).lower() in _ENLACES
        for token in tokens[numericos[0]:]
    )


def _cierra_oracion(anterior: str) -> bool:
    """¿El texto acumulado cierra una oración? (con o sin cita Vancouver)."""
    return bool(_FIN_DE_ORACION.search(anterior) or _FIN_CON_CITA.search(anterior))


def _es_linea_suelta(anterior: str, linea: str) -> bool:
    """¿`anterior` es una línea que va sola (un subtítulo que no se detectó)?

    Un subtítulo con el mismo cuerpo de letra que el texto y con un nombre no
    canónico ("Statistical analysis", "Sample size calculation") no lo cazan ni
    `detectar_seccion` ni `es_encabezado_por_formato`, y como no acaba en
    puntuación se pegaba a la primera frase del párrafo siguiente, que es la
    frase que mejor lo resume.

    Cuatro condiciones, todas necesarias porque cada una tapa un falso positivo
    medido con los casos del propio test de `_abre_parrafo`:
    - `anterior` tiene <= _MAX_PALABRAS_SUELTA palabras. `anterior` es el
      párrafo en construcción, así que esto solo puede ser cierto en su primera
      línea: una línea física de prosa lleva 10-15 palabras.
    - no acaba en signo de puntuación (una coma o un punto y coma es prosa).
    - no acaba en palabra de enlace ni en palabra con dígitos: "were assessed
      by", "as summarized in" y "the coefficient was" son prosa cortada, y "The
      mean difference was 0.31" o "We used gpt-5.4" (encontrados atacando esta
      misma regla: pocas palabras, mayúscula al principio y en la línea
      siguiente) son primera línea de párrafo, no subtítulos. Un subtítulo que
      acabe en número ("Experiment 2") se pierde a cambio, y no pasa nada: se
      queda pegado al párrafo que describe, que es lo de antes.
    - empieza por mayúscula y la línea siguiente también (un subtítulo abre el
      bloque; si la siguiente empieza en minúscula es su continuación).

    El compromiso que queda: una primera línea de párrafo de <=6 palabras que
    acabe en palabra plena y siga con mayúscula ("All patients received" |
    "Aricept ...") se parte de más. Es un párrafo en dos, no un dato perdido.
    """
    palabras = anterior.split()
    if not palabras or len(palabras) > _MAX_PALABRAS_SUELTA:
        return False
    if not anterior[-1:].isalnum():
        return False
    if palabras[-1].strip(_ADORNO_DE_CELDA).lower() in _ENLACES:
        return False
    if any(caracter.isdigit() for caracter in palabras[-1]):
        return False
    primera = re.search(r"[^\W_]", anterior)
    if primera is None or not primera.group(0).isupper():
        return False
    inicio = re.search(r"[^\W_]", linea)
    return inicio is not None and inicio.group(0).isupper()


def _abre_parrafo(anterior: str, linea: str) -> bool:
    """¿`linea` empieza un párrafo nuevo, o continúa el que acaba en `anterior`?

    pdfplumber devuelve líneas físicas, no párrafos, y hay que decidir dónde
    acaba uno con lo único que se tiene, que es el texto. La regla: la línea
    anterior cerró la oración (punto, interrogación, exclamación o dos puntos,
    con cierre de comillas o paréntesis si lo hay) Y la nueva arranca con
    mayúscula o cifra. Todo lo demás continúa el párrafo: una línea que acaba
    en "and" o en "hippocam-" no cierra nada, y minúscula tras punto es una
    abreviatura ("et al.", "e.g.") o un punto decimal.

    Se acepta cortar de más entre dos oraciones del mismo párrafo (ocurre
    cuando una oración termina justo a final de línea; con 10-15 palabras por
    línea, en torno al 7% de las oraciones): el empaquetador las vuelve a
    juntar, y una frontera entre oraciones es un buen sitio para cortar. Lo
    que NO se acepta es cortar de menos, que es lo que hacía el troceo por
    líneas: "between 2018 and" | "2021." y "542" | "pg/mL" partidos, y
    "hippocam-" | "pal" indexado como dos palabras que no existen.

    La regla del párrafo NO se aplica a lo que no es prosa, y por eso se miran
    dos cosas antes de la puntuación: una fila de tabla nunca se pega
    (ni como línea nueva ni como línea anterior; ver `_parece_fila_de_tabla`,
    regresión medida el 4 sep 2026: las tablas de un PDF salían con todas las
    filas en una línea) y una línea que va sola tampoco (`_es_linea_suelta`,
    el subtítulo no detectado que se comía la primera frase del párrafo).
    """
    if _VINETA.match(linea):
        return True
    # Una fila de tabla es un párrafo propio: nada se le pega delante ni detrás.
    if _parece_fila_de_tabla(linea):
        return True
    # El rótulo de tabla solo abre párrafo si la anterior NO sigue a mitad de
    # oración: si sigue, "Table 1." es una referencia dentro de la frase ("as
    # summarized in" | "Table 1. The groups did not differ.") y su punto es el
    # final de ESA frase, no el de un rótulo.
    if _ROTULO.match(linea) and (
        _cierra_oracion(anterior) or _parece_fila_de_tabla(anterior)
    ):
        return True
    if _GUION_FINAL.search(anterior):
        return False
    if _parece_fila_de_tabla(anterior):
        return True
    if not _cierra_oracion(anterior):
        return _es_linea_suelta(anterior, linea)
    bajo = anterior.lower()
    if any(bajo.endswith(abreviatura) for abreviatura in _ABREVIATURAS):
        return False
    # Primer carácter con contenido, saltando comillas y paréntesis de apertura.
    inicio = re.search(r"[^\W_]", linea)
    if inicio is None:
        return True
    return inicio.group(0).isupper() or inicio.group(0).isdigit()


def _unir_lineas(anterior: str, linea: str) -> str:
    """Pega `linea` al final de `anterior`, deshaciendo el corte de palabra.

    "hippocam-" + "pal" → "hippocampal": el guion lo puso el maquetador y sin
    quitarlo la palabra no existe en el índice (medido con el parser: tras
    indexar un párrafo con "hippocam-" a final de línea, "hippocampal" no
    aparecía en ningún chunk). Se quita solo cuando lo precede una letra y la
    continuación es minúscula; con mayúscula o cifra detrás ("anti-" +
    "Alzheimer", "COVID-" + "19") el guion es parte del término y se conserva.

    Y se conserva TAMBIÉN cuando la parte izquierda es un prefijo conocido
    (_PREFIJOS_CON_GUION) o tiene <= _MAX_IZQUIERDA_CON_GUION caracteres: un
    compuesto partido justo en su propio guion perdía el guion y "anti-" +
    "inflammatory" quedaba indexado como "antiinflammatory", "beta-" +
    "amyloid" como "betaamyloid". Son términos centrales del corpus, así que la
    forma sin guion no la encuentra nadie: dejaban de existir en el índice.

    El compromiso: una palabra cuya primera mitad sea corta y no esté en la
    lista ("hip-" + "pocampal", "pa-" + "tients") se queda con un guion que no
    le toca. Se elige ese lado porque un guion de más deja el término buscable
    por sus dos mitades, mientras que un guion de menos inventa una palabra que
    no existe; y el caso que motivó la regla, "hippocam-" + "pal" (izquierda de
    8 caracteres), sigue uniéndose en "hippocampal".
    """
    if _CORTE_DE_PALABRA.search(anterior) and linea[:1].islower():
        izquierda = re.search(r"([^\W\d_]+)[-\u2010\u00ad]$", anterior)
        parte = izquierda.group(1).lower() if izquierda else ""
        if parte in _PREFIJOS_CON_GUION or len(parte) <= _MAX_IZQUIERDA_CON_GUION:
            return anterior + linea
        return anterior[:-1] + linea
    if _GUION_FINAL.search(anterior):
        return anterior + linea
    return f"{anterior} {linea}"


def _decode_bytes(raw: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------
def _parse_pdf(
    path: Path, file_name: str, skip_references: bool = True
) -> tuple[list[dict], int, int]:
    """PDF con conciencia de documento: sección vigente y metadatos de la obra.

    Devuelve (chunks, páginas, líneas_de_bibliografía_descartadas).

    La sección se arrastra desde el último encabezado, y un encabezado se
    reconoce de dos maneras: por su nombre, cuando es una sección de artículo
    conocida (Métodos, Resultados), y por su MAQUETA, cuando no lo es. La
    segunda hace falta porque en un documento que no es un paper los
    encabezados se llaman "Composición del mazo" o "Por qué elegirnos": sin
    detectarlos, la primera sección reconocida se arrastraría hasta el final y
    el fragmento de la página 4 acabaría citado como "sección: Introducción",
    que es peor que no decir nada.

    La bibliografía se descarta por defecto: son títulos de trabajos ajenos que
    matchean con casi cualquier consulta sin ser evidencia de nada, y ocupan
    una parte nada despreciable de lo que se paga por embeber.
    """
    import pdfplumber

    from app.ingest import paper as paper_mod

    # (párrafo, páginas que abarca, sección)
    paras: list[tuple[str, list[int], str]] = []
    meta = paper_mod.PaperMeta()
    descartados = 0
    seccion = ""
    canonica = ""

    # Primera pasada: texto y formato de cada página. Hace falta el documento
    # entero antes de empezar, por dos razones: las cabeceras y pies se
    # detectan por repetirse entre páginas, y el tamaño del texto corrido (con
    # el que se reconocen los encabezados) solo se sabe mirándolo todo.
    paginas: list[tuple[int, list[str]]] = []
    formato: list[list[tuple[str, float, bool]]] = []
    chars: list[dict] = []
    cabecera_texto: list[str] = []

    with pdfplumber.open(path) as pdf:
        page_count = len(pdf.pages)
        for page_no, page in enumerate(pdf.pages, start=1):
            try:
                text = page.extract_text() or ""
            except Exception as exc:
                logger.warning(
                    "%s pág. %d: fallo extrayendo texto (%s); se omite.",
                    file_name, page_no, exc,
                )
                continue
            try:
                chars_pagina = page.chars or []
            except Exception:
                chars_pagina = []
            if page_no == 1:
                chars = chars_pagina
            if page_no <= 2:
                cabecera_texto.append(text)
            paginas.append((page_no, text.splitlines()))
            formato.append(paper_mod.lineas_con_formato(chars_pagina))

    if chars:
        meta = paper_mod.extraer_metadatos(chars, "\n".join(cabecera_texto))

    repetidas = paper_mod.lineas_repetidas([lineas for _, lineas in paginas])
    cuerpo = paper_mod.tamano_de_cuerpo(formato)
    # Formato por texto de línea, para consultarlo mientras se recorre el texto
    # plano (que es lo que da los párrafos). Si una línea se repite con formatos
    # distintos gana el primero, que es suficiente para decidir si es titular.
    formato_por_texto: dict[str, tuple[float, bool]] = {}
    for lineas_pag in formato:
        for texto_linea, tamano, negrita in lineas_pag:
            formato_por_texto.setdefault(texto_linea.strip(), (tamano, negrita))

    # Segunda pasada: secciones y párrafos. Los encabezados se detectan línea a
    # línea, porque un encabezado suele ser su propia línea corta. El resto de
    # líneas se van UNIENDO en párrafos (ver _abre_parrafo): pdfplumber da
    # líneas físicas, y trocear por ellas (lo que se hacía hasta sep 2026)
    # partía oraciones, separaba cifras de sus unidades y dejaba las palabras
    # cortadas con guion como dos mitades que no existen.
    abierto = ""                     # párrafo en construcción
    abierto_paginas: list[int] = []  # páginas que toca (puede cruzar de página)

    def _cerrar_parrafo() -> None:
        nonlocal abierto, abierto_paginas
        texto = abierto.strip()
        if texto:
            for pieza in _split_long_paragraph(texto):
                paras.append((pieza, list(abierto_paginas), seccion))
        abierto = ""
        abierto_paginas = []

    for page_no, lineas in paginas:
        utiles = [l for l in lineas if l.strip()]
        for i, linea in enumerate(utiles):
            if paper_mod.es_ruido_de_pagina(linea):
                continue
            if (
                paper_mod.en_borde(i, len(utiles))
                and paper_mod.normalizar(linea) in repetidas
            ):
                continue
            detectada = paper_mod.detectar_seccion(linea)
            if detectada is None:
                # No es una sección con nombre conocido, pero puede ser un
                # encabezado igualmente: si lo es, RESETEA la sección en vez de
                # dejar que la anterior se arrastre por un contenido que no
                # describe.
                tamano, negrita = formato_por_texto.get(linea.strip(), (0.0, False))
                if paper_mod.es_encabezado_por_formato(
                    linea, tamano, cuerpo, negrita
                ):
                    # El párrafo abierto se cierra ANTES de cambiar la sección:
                    # pertenece a la anterior.
                    _cerrar_parrafo()
                    canonica = ""
                    seccion = linea.strip()
                    continue
            else:
                _cerrar_parrafo()
                canonica = detectada
                seccion = linea.strip()
                continue
            if skip_references and canonica == paper_mod.REFERENCIAS:
                descartados += 1
                continue
            linea = linea.strip()
            if abierto and _abre_parrafo(abierto, linea):
                _cerrar_parrafo()
            abierto = _unir_lineas(abierto, linea) if abierto else linea
            if page_no not in abierto_paginas:
                abierto_paginas.append(page_no)
    _cerrar_parrafo()

    # Se empaqueta por tramos de sección (ver _agrupar_por_seccion), con el
    # solape de _pack_paragraphs, y cada chunk lleva delante el título de la
    # obra y su sección: un fragmento de Resultados tiene que decir que lo es
    # también en el texto que se embebe, no solo en el payload.
    chunks: list[dict] = []
    for sec, grupo in _agrupar_por_seccion(paras):
        for paquete in _pack_paragraphs(grupo):
            cuerpo_chunk = "\n\n".join(t for t, _ in paquete).strip()
            if not cuerpo_chunk:
                continue
            pags = sorted({pg for _, pgs in paquete for pg in pgs})
            chunks.append(
                _base_chunk(
                    file_name, _con_contexto(cuerpo_chunk, meta.titulo, sec),
                    pags[0], pags, "text", section=sec, meta=meta,
                )
            )

    return chunks, page_count, descartados


# ---------------------------------------------------------------------------
# Tabulares (XLSX / CSV)
# ---------------------------------------------------------------------------
def _cell_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _looks_numeric(s: str) -> bool:
    return bool(re.fullmatch(r"[-+]?[\d.,%$ ]+", s))


def _detect_header(rows: list[tuple[int, list[str]]]) -> tuple[list[str] | None, int]:
    """(header, índice en `rows` de la fila de header) o (None, -1).

    Header = la primera fila con ≥2 celdas no vacías, si ≥60% de esas celdas
    son textuales (no numéricas). Si esa primera fila 'ancha' es numérica,
    se asume que no hay encabezado.
    """
    for idx, (_row_no, cells) in enumerate(rows[:10]):
        non_empty = [c for c in cells if c]
        if len(non_empty) < 2:
            continue
        texty = [c for c in non_empty if not _looks_numeric(c)]
        if len(texty) / len(non_empty) >= 0.6:
            return cells, idx
        return None, -1
    return None, -1


def _rows_to_chunks(
    rows: list[tuple[int, list[str]]],
    file_name: str,
    sheet_label: str | None = None,
) -> list[dict]:
    """Un chunk por fila de datos ("Campo: valor"). page = número de fila."""
    if not rows:
        return []
    header, header_idx = _detect_header(rows)
    data_rows = rows[header_idx + 1:] if header is not None else rows

    def field_name(j: int) -> str:
        if header is not None and j < len(header) and header[j]:
            return header[j]
        return f"Columna {j + 1}"

    chunks: list[dict] = []
    for row_no, cells in data_rows:
        lines = [
            f"{field_name(j)}: {val}"
            for j, val in enumerate(cells)
            if val
        ]
        if not lines:
            continue
        if sheet_label:
            lines.insert(0, f"Hoja: {sheet_label}, fila {row_no}")
        text = "\n".join(lines).strip()
        if not text:
            continue
        chunks.append(
            _base_chunk(
                file_name, text, row_no, [row_no], "table", source_row=row_no
            )
        )
    return chunks


def _parse_xlsx(path: Path, file_name: str) -> tuple[list[dict], int]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets_with_data: list[tuple[str, list[tuple[int, list[str]]]]] = []
        for ws in wb.worksheets:
            rows: list[tuple[int, list[str]]] = []
            for row_no, row in enumerate(ws.iter_rows(values_only=True), start=1):
                cells = [_cell_str(v) for v in row]
                if any(cells):
                    rows.append((row_no, cells))
                # Corta temprano cuando el tope ya está garantizadamente
                # excedido (+10: margen por la posible fila de header).
                if len(rows) > MAX_CHUNKS + 10:
                    break
            if rows:
                sheets_with_data.append((ws.title, rows))
    finally:
        wb.close()

    multi = len(sheets_with_data) > 1
    chunks: list[dict] = []
    for title, rows in sheets_with_data:
        chunks.extend(
            _rows_to_chunks(rows, file_name, sheet_label=title if multi else None)
        )
    return chunks, len(chunks)


def _parse_csv(path: Path, file_name: str) -> tuple[list[dict], int]:
    text = _decode_bytes(path.read_bytes())
    sample = text[:8192]
    try:
        dialect: csv.Dialect | type[csv.Dialect] = csv.Sniffer().sniff(
            sample, delimiters=",;\t|"
        )
    except csv.Error:
        dialect = csv.excel

    rows: list[tuple[int, list[str]]] = []
    reader = csv.reader(io.StringIO(text), dialect)
    for row_no, row in enumerate(reader, start=1):
        cells = [c.strip() for c in row]
        if any(cells):
            rows.append((row_no, cells))
        if len(rows) > MAX_CHUNKS + 10:  # ver comentario en _parse_xlsx
            break

    chunks = _rows_to_chunks(rows, file_name)
    return chunks, len(chunks)


# ---------------------------------------------------------------------------
# TXT / MD
# ---------------------------------------------------------------------------
def _parse_text(path: Path, file_name: str) -> tuple[list[dict], int]:
    text = _decode_bytes(path.read_bytes())
    paras = [(p, 0) for p in _split_paragraphs(text)]

    chunks: list[dict] = []
    for i, group in enumerate(_pack_paragraphs(paras), start=1):
        chunk_text = "\n\n".join(p for p, _ in group).strip()
        if not chunk_text:
            continue
        # page = índice de chunk (1-based): no hay páginas reales.
        chunks.append(_base_chunk(file_name, chunk_text, i, [i], "text"))
    return chunks, len(chunks)


# ---------------------------------------------------------------------------
# Word (.docx)
# ---------------------------------------------------------------------------
def _es_titulo(parrafo) -> bool:
    """¿El párrafo es un encabezado? Por estilo, que es lo fiable en Word."""
    nombre = (getattr(parrafo.style, "name", "") or "").lower()
    return nombre.startswith(("heading", "título", "titulo", "subtitle", "subtítulo"))


def _filas_de_tabla(tabla) -> list[list[str]]:
    """Filas de una tabla de Word como listas de celdas, UNA POR COLUMNA de la
    cuadrícula.

    Devuelve celdas y no la línea ya montada porque la cabecera (ver
    `_cabecera_de_tabla`) se decide contando celdas efectivas, y contar '|' en
    un texto es adivinar.

    Dos formas de perder columnas, las dos medidas con documentos reales:

    1. Deduplicar por TEXTO creyendo que dos celdas iguales son una combinada.
       Una tabla de basales con dos grupos de la misma edad media, "Edad | 72.4
       (6.1) | 72.4 (6.1) | 74.0 (5.8) | 0.31", salía como "Edad | 72.4 (6.1) |
       74.0 (5.8) | 0.31": 74.0 pasaba a leerse bajo MCI y 0.31 bajo AD.
    2. Colapsar la combinada de verdad a UNA posición. Es lo que hacía la
       deduplicación por identidad del `_tc` hasta el 4 sep 2026: en una tabla
       de cabecera ["Grupo","Basal","Final","p"], la fila con la celda de
       "Basal"+"Final" combinada salía como "AD | n=40 (both visits) | 0.03",
       tres columnas contra cuatro, y 0.03 se leía como "Final".

    La regla que arregla las dos: se recorre la cuadrícula y una celda que
    abarca k columnas (su `w:gridSpan`, que python-docx expone como
    `celda._tc.grid_span`) ocupa k posiciones, con el TEXTO EN LA PRIMERA y un
    marcador vacío en las k-1 siguientes. Así la posición de cada valor es su
    columna, que es lo único que hace falta para leerlo bajo su cabecera.

    Marcador vacío y no texto repetido a propósito: repetirlo pondría el valor
    bajo una cabecera en la que no se midió ("0.03" bajo "Final" es justo el
    error que se está corrigiendo), y una fila de título combinada a todo el
    ancho se repetiría cuatro veces en cada bloque de la tabla.

    Las combinadas VERTICALES (vMerge) sí repiten su texto en cada fila, y eso
    se deja como está: es lo que python-docx da en `fila.cells` y lo que permite
    leer una fila sola sin buscar el rótulo de grupo filas más arriba.

    Las celdas vacías se conservan en su sitio y solo se recorta el final de la
    fila (en `_fila_a_texto`): el hueco de la esquina superior izquierda de una
    tabla de basales es una columna, y quitarlo desplazaría la cabecera
    respecto a los datos.
    """
    filas: list[list[str]] = []
    for fila in tabla.rows:
        # Word permite que una fila empiece más allá de la primera columna;
        # esos huecos son columnas y hay que reponerlos (python-docx los omite
        # de `fila.cells` y los cuenta en `grid_cols_before`).
        celdas: list[str] = [""] * max(0, getattr(fila, "grid_cols_before", 0) or 0)
        posiciones = list(fila.cells)
        i = 0
        while i < len(posiciones):
            celda = posiciones[i]
            ancho = getattr(getattr(celda, "_tc", None), "grid_span", 1) or 1
            ancho = max(1, min(ancho, len(posiciones) - i))
            celdas.append(re.sub(r"\s+", " ", celda.text).strip())
            celdas.extend([""] * (ancho - 1))
            i += ancho
        if not any(celdas):
            continue
        filas.append(celdas)
    return filas


def _fila_a_texto(celdas: list[str]) -> str:
    """Fila como línea de texto, celdas separadas por ' | '.

    Se separan con ' | ' para que el modelo pueda leer la fila entera; las
    tablas de un documento clínico suelen llevar los datos que a nadie le sirve
    perder. Del final se recortan las celdas vacías, que no llevan ningún valor
    cuya columna se pueda confundir: una combinada que llega hasta la última
    columna pierde así su extensión visual, pero ningún dato cambia de sitio.
    """
    return " | ".join(celdas).rstrip(" |")


def _cabecera_de_tabla(filas: list[list[str]]) -> int:
    """Cuántas filas iniciales son cabecera y hay que repetir en cada bloque.

    Normalmente una. Pero en las tablas clínicas de Word la fila 0 es a menudo
    un TÍTULO combinado a todo el ancho ("Table 1. Baseline characteristics") y
    la cabecera real es la fila 1; tomando ciegamente la fila 0 (lo que se hacía
    hasta el 4 sep 2026), los bloques 2..N repetían el título y NO llevaban los
    nombres de columna, que es exactamente lo que se quería evitar: medido con
    una tabla de 200 filas, las cinco partes salían encabezadas por "Table 1.
    Baseline characteristics" y sin "ID | Grupo | Basal | Final".

    Criterio: la fila 0 tiene UNA SOLA celda efectiva (no vacía) en una
    cuadrícula de dos o más columnas. Sirve para el título combinado y también
    para la cabecera de dos pisos ("" | "Grupo" abarcando dos columnas, y debajo
    "Variable" | "Control" | "AD"): en los dos casos la fila 0 no nombra las
    columnas, así que la que las nombra es la 1 y las dos tienen que viajar
    juntas. Se exige que queden filas de datos (>=3 en total) para no dejar un
    bloque que sea solo cabecera.
    """
    if len(filas) >= 3 and len(filas[0]) >= 2 and sum(1 for c in filas[0] if c) == 1:
        return 2
    return 1


def _tabla_en_bloques(filas: list[list[str]]) -> list[str]:
    """Reparte las filas de una tabla en bloques de ~_TARGET_TOKENS, cada uno
    encabezado por las filas de cabecera.

    Antes la tabla era un único chunk recortado a _MAX_CHUNK_CHARS: en una
    tabla larga las filas del final desaparecían sin aviso. Y en un bloque que
    no sea el primero "74.0 (5.8)" no significa nada sin la fila "Control |
    MCI | AD | p" que le da columna, así que la cabecera se repite en todos.
    Cuántas filas son cabecera lo decide `_cabecera_de_tabla`: no siempre es la
    fila 0, porque en Word suele ser un título combinado a todo el ancho.
    """
    if len(filas) < 2:
        return [_fila_a_texto(f) for f in filas]
    corte = _cabecera_de_tabla(filas)
    cabecera = "\n".join(_fila_a_texto(f) for f in filas[:corte])
    cuerpo = [_fila_a_texto(f) for f in filas[corte:]]
    presupuesto = max(_TARGET_TOKENS - _est_tokens(cabecera), 1)
    bloques: list[str] = []
    actual: list[str] = []
    actual_tok = 0
    for fila in cuerpo:
        tok = _est_tokens(fila)
        if actual and actual_tok + tok > presupuesto:
            bloques.append("\n".join([cabecera, *actual]))
            actual, actual_tok = [], 0
        actual.append(fila)
        actual_tok += tok
    if actual:
        bloques.append("\n".join([cabecera, *actual]))
    return bloques


def _parse_docx(path: Path, file_name: str) -> tuple[list[dict], int]:
    """Word: párrafos agrupados por sección, más las tablas del documento.

    Word no tiene páginas: el salto de página lo calcula el visor al
    renderizar, así que el localizador de cita es la sección (el encabezado
    vigente) y, a falta de encabezados, el número de fragmento.
    """
    import docx
    from docx.table import Table

    documento = docx.Document(str(path))

    # (texto, sin localizador, sección) en el orden del documento.
    bloques: list[tuple[str, None, str]] = []
    # (filas de celdas, sección, rótulo) por tabla, en el orden del documento.
    tablas: list[tuple[list[list[str]], str, str]] = []
    seccion = ""
    ultimo_parrafo = ""

    # Se recorre el cuerpo en orden de documento (iter_inner_content, python-docx
    # >= 1.0) y no `paragraphs` + `tables` por separado: es lo que permite que
    # una tabla herede la sección bajo la que aparece (una tabla de basales va
    # en Resultados) y que se lleve el rótulo "Tabla 1. ..." que la precede.
    for elemento in documento.iter_inner_content():
        if isinstance(elemento, Table):
            filas = _filas_de_tabla(elemento)
            if filas:
                rotulo = ultimo_parrafo if _ROTULO_TABLA.match(ultimo_parrafo) else ""
                tablas.append((filas, seccion, rotulo))
            # Un rótulo describe UNA tabla: sin este reset, dos tablas
            # consecutivas heredaban las dos "Table 1. ..." (medido el 4 sep
            # 2026) y la segunda quedaba citada como la primera.
            ultimo_parrafo = ""
            continue
        texto = elemento.text.strip()
        if not texto:
            continue
        if _es_titulo(elemento):
            seccion = texto
            # El título también se indexa: es la mejor pista de qué viene.
            bloques.append((texto, None, seccion))
            # Un rótulo de tabla es el párrafo INMEDIATAMENTE anterior a ella;
            # si en medio hay un encabezado, ya no describe la tabla.
            ultimo_parrafo = ""
            continue
        for pieza in _split_long_paragraph(texto):
            bloques.append((pieza, None, seccion))
        ultimo_parrafo = texto

    chunks: list[dict] = []
    indice = 0

    # Empaquetado por tramos de sección con el solape de _pack_paragraphs, y la
    # sección vigente delante de cada chunk (salvo el primero de la sección,
    # que ya empieza por el encabezado indexado como bloque).
    for sec, grupo in _agrupar_por_seccion(bloques):
        for paquete in _pack_paragraphs(grupo):
            cuerpo = "\n\n".join(t for t, _ in paquete).strip()
            if not cuerpo:
                continue
            indice += 1
            chunks.append(
                _base_chunk(
                    file_name, _con_contexto(cuerpo, sec), indice, [indice],
                    "text", section=sec,
                )
            )

    # Las tablas se numeran aparte (tabla 1, tabla 2...): es como las busca
    # quien abre el documento, y no comparten numeración con los párrafos. Una
    # tabla larga sale en varios chunks que citan el MISMO número de tabla, y
    # cada uno lleva la cabecera y la sección para leerse por sí solo.
    for numero, (filas, sec, rotulo) in enumerate(tablas, start=1):
        partes = _tabla_en_bloques(filas)
        for parte, texto in enumerate(partes, start=1):
            chunk = _base_chunk(
                file_name, _con_contexto(texto, sec, rotulo), numero, [numero],
                "table", section=sec,
            )
            if len(partes) > 1:
                chunk["metadata"].update({"table_part": parte, "table_parts": len(partes)})
            chunks.append(chunk)

    return chunks, len(chunks)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".xlsx", ".csv", ".txt", ".md"}


def parse_generic(
    path: Path, file_name: str, skip_references: bool = True
) -> tuple[list[dict], int]:
    """Parsea un documento → (chunks, pages).

    `pages`: nº de páginas para PDF; nº de filas/chunks para el resto (es lo
    que se muestra como "pages" en GET /documents).

    `skip_references`: descarta la bibliografía de un PDF (default). Son
    títulos de trabajos ajenos: matchean con cualquier consulta, no son
    evidencia de nada y se pagan igual al embeberlos. Ponerlo en False solo
    tiene sentido si lo que se quiere consultar ES la bibliografía.

    Levanta ValueError si la extensión no está soportada, si no se extrae
    texto alguno, o si se supera MAX_CHUNKS.
    """
    ext = path.suffix.lower()
    descartados = 0
    if ext == ".pdf":
        chunks, pages, descartados = _parse_pdf(
            path, file_name, skip_references=skip_references
        )
        if descartados:
            logger.info(
                "%s: %d líneas de bibliografía descartadas.", file_name, descartados
            )
    elif ext == ".docx":
        chunks, pages = _parse_docx(path, file_name)
    elif ext == ".doc":
        raise ValueError(
            "El formato .doc (Word 97-2003) no se puede leer. Abre el archivo "
            "en Word y guárdalo como .docx, o expórtalo a PDF."
        )
    elif ext == ".xlsx":
        chunks, pages = _parse_xlsx(path, file_name)
    elif ext == ".csv":
        chunks, pages = _parse_csv(path, file_name)
    elif ext in (".txt", ".md"):
        chunks, pages = _parse_text(path, file_name)
    else:
        raise ValueError(f"Extensión no soportada: {ext}")

    # Saneo final: sin texto → fuera (defensa extra; ya se filtra antes).
    chunks = [c for c in chunks if c["text"].strip()]

    # Idioma del documento entero, no por fragmento: un artículo está escrito
    # en un idioma, y decidirlo sobre todo el texto es mucho más fiable que
    # sobre un párrafo corto. Si no queda claro se deja vacío.
    idioma = detectar_idioma("\n".join(c["text"] for c in chunks[:40]))
    for chunk in chunks:
        chunk["language"] = idioma

    if not chunks:
        raise ValueError(
            f"'{file_name}' no contiene texto extraíble (¿PDF escaneado sin OCR "
            f"o archivo vacío?)"
        )
    if len(chunks) > MAX_CHUNKS:
        raise ValueError(
            f"'{file_name}' genera {len(chunks)} chunks; el máximo permitido es "
            f"{MAX_CHUNKS}. Divide el documento en archivos más pequeños."
        )
    return chunks, pages
